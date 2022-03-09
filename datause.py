#!/user/bin/python
# -*- coding:utf-8 -*-
# 作者：realityerror
# 创建：2022-03-06
# 更新：2022-03-06
# 用意：用于搭建简单的xlsx数据库,并进行操作

from ast import Str
import init
import openpyxl as use


class database():#选择文件和其中的表格

    #初始化数据对象
    def __init__(self,filename,sheetname):
        self.file = filename
        self.xlsx = use.load_workbook(self.file)
        self.user = self.xlsx[sheetname]


    #打印全部内容
    def print_all(self):
        for j in self.user.rows:    # we.rows 获取每一行数据
            for n in j:
                print(n.value, end="\t")   # n.value 获取单元格的值
            print( )

    # #行的最大值
    # def max_x(self):
    #     num_x = self.user.max_column
    #     return num_x

    # #列的最大值
    # def max_y(self):
    #     num_y = self.user.max_row
    #     return num_y

    #找到具体的某个值
    def find_one(self,x,y):
        cell_value = self.user.cell(row=x, column=y).value
        return cell_value

    #获取某行的数值
    def find_R(self,row):
        columns = self.user.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.user.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 获取某列的所有值
    def find_C(self, column):
        rows = self.user.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.user.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 设置某个单元格的值
    def set_one(self, row, colunm, cellvalue):
        try:
            self.user.cell(row=row, column=colunm).value = cellvalue
            self.xlsx.save(self.file)
        except:
            self.user.cell(row=row, column=colunm).value = "writefail"
            self.xlsx.save(self.file)

    # 根据数据库自动生成字典
    def zidian_ro(self):
        r = self.user.max_row #y
        l = self.user.max_column#x
        dict = {}
        for i in range(1, r+1):
            title  =  self.user.cell(i, 1).value
            #print(title,1)
            for j in range(1,l+1):               
                value =  self.user.cell(i, j).value
                #print(value,2)
                dict[title] = value
        #print(dict)
        return dict 


#xlsx_user = use.load_workbook(filename)
#
#user_num = user.max_column
#print(user_num-1)