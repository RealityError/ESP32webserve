#!/user/bin/python
# -*- coding:utf-8 -*-
# 作者：realityerror
# 创建：2022-03-06
# 更新：2022-03-06
# 用意：用于搭建简单的xlsx数据库,并进行操作

import init
import openpyxl as use


class database(filename,sheetname):
    def __init__(self,filename,sheetname):
        self.xlsx = use.load_workbook(filename)
        self.user = self.xlsx[sheetname]

    def max_x(self):
        num_x = self.user.max_column
        return num_x-1
    def max_y(self):
        num_y = self.user.max_column
        return num_x-1

#xlsx_user = use.load_workbook(filename)
#
#user_num = user.max_column
#print(user_num-1)